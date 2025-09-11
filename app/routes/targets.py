from __future__ import annotations

import io, os, csv
from collections import Counter, defaultdict
from typing import Optional, Iterable

from flask import Blueprint, render_template, request, redirect, url_for, flash
from sqlalchemy import select, text, or_

from ..extensions import db
from .. import extensions as ext
from ..models import TargetList, TargetListEntry, Program, Campaign, Brand

targets_bp = Blueprint("targets", __name__, template_folder="../templates/targets")

# -----------------------------
# CSV/XLSX parsing (no pandas)
# -----------------------------
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

def _read_csv_dicts(file_bytes: bytes):
    f = io.StringIO(file_bytes.decode('utf-8', errors='ignore'))
    reader = csv.DictReader(f)
    return list(reader)

def _read_xlsx_dicts(file_bytes: bytes):
    if not load_workbook:
        raise RuntimeError("openpyxl not installed")
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        rec = {}
        for i, v in enumerate(r):
            key = headers[i] if i < len(headers) else f"col{i+1}"
            rec[key] = "" if v is None else str(v)
        out.append(rec)
    return out

def _read_rows(file_storage):
    content = file_storage.read()
    file_storage.seek(0)
    name = (file_storage.filename or "").lower()
    if name.endswith(('.xlsx', '.xls')):
        return _read_xlsx_dicts(content)
    else:
        return _read_csv_dicts(content)

def _pick_npi_column(rows):
    if not rows:
        return None
    headers = list(rows[0].keys())
    lower = [h.lower().strip() for h in headers]
    candidates = ["npi", "npi_id", "npi number", "npi_number"]
    for cand in candidates:
        if cand in lower:
            return headers[lower.index(cand)]
    best = None; best_score = -1
    for h in headers:
        vals = [("".join([c for c in (row.get(h) or "") if c.isdigit()])) for row in rows[:200]]
        score = sum(1 for v in vals if len(v) == 10)
        if score > best_score:
            best_score = score; best = h
    return best or headers[0]

def _extract_clean_npi(raw: str):
    return "".join(ch for ch in str(raw or "").strip() if ch.isdigit())

# -----------------------------
# Program search helper
# -----------------------------
def _query_programs(q: Optional[str]):
    stmt = select(Program).order_by(Program.created_at.desc())
    if q:
        like = f"%{q}%"
        stmt = (
            select(Program)
            .join(Campaign, Program.campaign_id == Campaign.id, isouter=True)
            .join(Brand, Campaign.brand_id == Brand.id, isouter=True)
            .where(
                or_(
                    Program.name.ilike(like),
                    Program.platform.ilike(like),
                    Program.type.ilike(like),
                    Campaign.name.ilike(like),
                    Brand.name.ilike(like),
                )
            )
            .order_by(Program.created_at.desc())
        )
    return db.session.execute(stmt).scalars().all()

# -----------------------------
# Data summary (from JSON extras)
# -----------------------------
SUMMARY_CANDIDATES = [("Specialty", 8), ("Segment", 8), ("Tier", 5)]
NUMERIC_CANDIDATES = ["ActivityScore", "Score", "Rank"]

def _summarize_entries(tid: int, limit_rows: int = 20000):
    rows = db.session.execute(
        select(TargetListEntry.extra, TargetListEntry.npi).where(TargetListEntry.target_list_id == tid).limit(limit_rows)
    ).all()
    facet_counts = {k: Counter() for k, _ in SUMMARY_CANDIDATES}
    numeric_values = defaultdict(list)

    for extra, _npi in rows:
        if not extra or not isinstance(extra, dict):
            continue
        for key, topn in SUMMARY_CANDIDATES:
            if key in extra and extra[key]:
                facet_counts[key][str(extra[key])] += 1
        for nk in NUMERIC_CANDIDATES:
            if nk in extra:
                try:
                    numeric_values[nk].append(float(str(extra[nk]).strip()))
                except Exception:
                    pass

    facets = {}
    for key, topn in SUMMARY_CANDIDATES:
        if facet_counts[key]:
            facets[key] = facet_counts[key].most_common(topn)

    numerics = {}
    for nk, vals in numeric_values.items():
        if vals:
            vals_sorted = sorted(vals)
            n = len(vals_sorted)
            def pct(p):
                i = max(0, min(n-1, int(round((p/100.0)*(n-1)))) )
                return vals_sorted[i]
            numerics[nk] = {"count": n, "min": vals_sorted[0], "p50": pct(50), "p90": pct(90), "max": vals_sorted[-1]}

    sample = db.session.execute(
        select(TargetListEntry.npi, TargetListEntry.extra).where(TargetListEntry.target_list_id == tid).limit(25)
    ).all()

    return facets, numerics, sample

# -----------------------------
# Matching helpers (PG read-only + Python set intersection)
# -----------------------------

def _env_true(name: str, default: str = "1") -> bool:
    v = os.getenv(name, default)
    return str(v).lower() in {"1", "true", "yes", "y", "on"}

def _stream_network_npis_from_pg() -> Iterable[str]:
    """
    Stream DISTINCT NPIs from Postgres using a read-only SQLAlchemy engine and
    the SQL snippet provided by ext.ro_sql_network_npi. This function only READS
    from Postgres; it does not create temp tables or write anything to PG.
    """
    if not getattr(ext, "ro_engine", None) or not getattr(ext, "ro_sql_network_npi", None):
        return iter(())  # nothing configured
    # Ensure the inner SQL selects a column named "npi"
    outer_sql = "SELECT DISTINCT net.npi FROM (" + ext.ro_sql_network_npi + ") AS net"
    def _gen():
        with ext.ro_engine.connect() as conn:
            try:
                conn.exec_driver_sql("SET TRANSACTION READ ONLY;")
            except Exception:
                pass
            result = conn.exec_driver_sql(outer_sql)
            # fetch in chunks to keep memory low
            while True:
                rows = result.fetchmany(50000)
                if not rows:
                    break
                for (npi,) in rows:
                    # Cast to str and strip just in case
                    yield str(npi).strip()
    return _gen()

def compute_network_match_count(target_list_id: int) -> int:
    """
    Compute network overlap for a target list.

    - If cache strategy is 'startup_snapshot' or 'manual', uses local SQLite table `network_npis`.
    - Otherwise, streams NPIs from Postgres (read-only) and intersects with the SQLite target list NPIs in Python.
    - Never raises: returns 0 on error and flashes a warning.
    """
    if not _env_true("MATCH_NETWORK_ENABLED", "1"):
        return 0
    try:
        strategy = getattr(ext, "network_cache_strategy", "live")
        use_cache = strategy in {"startup_snapshot", "manual"}

        if use_cache:
            if hasattr(ext, "ensure_local_network_table"):
                ext.ensure_local_network_table()
            res = db.session.execute(text("""
                SELECT COUNT(*) FROM (
                  SELECT npi FROM network_npis
                  WHERE npi IN (SELECT e.npi FROM target_list_entries e WHERE e.target_list_id = :tid)
                ) sub
            """), {"tid": target_list_id}).scalar_one()
            return int(res or 0)

        # ---- LIVE MATCH (read Postgres, compare in Python) ----
        # 1) pull target list NPIs from SQLite
        tl_npis = set(
            db.session.execute(
                select(TargetListEntry.npi).where(TargetListEntry.target_list_id == target_list_id)
            ).scalars()
        )
        if not tl_npis:
            return 0

        # 2) stream Postgres network NPIs and intersect
        matched = 0
        for net_npi in _stream_network_npis_from_pg():
            if net_npi in tl_npis:
                matched += 1

        return matched

    except Exception as e:
        try:
            flash(f"Network match skipped: {type(e).__name__}: {e}", "warning")
        except Exception:
            pass
        return 0

# -----------------------------
# Routes
# -----------------------------
@targets_bp.route("/target-lists", methods=["GET", "POST"])
def target_lists():
    # Ensure optional columns exist on SQLite (no-op on Postgres)
    try:
        if hasattr(ext, "ensure_mvp_schema_sqlite"):
            ext.ensure_mvp_schema_sqlite()
    except Exception:
        pass

    if request.method == "POST":
        name = request.form.get("name", "").strip() or "Unnamed List"
        brand = request.form.get("brand") or None
        pharma = request.form.get("pharma") or None
        indication = request.form.get("indication") or None
        notes = request.form.get("notes") or None

        f = request.files.get("file")
        if not f or f.filename == "":
            flash("Upload a CSV/XLSX with an NPI column.", "warning")
            return redirect(url_for("targets.target_lists"))
        try:
            # Read & identify NPI column
            rows = _read_rows(f)
            if not rows:
                raise RuntimeError("File appears empty.")
            npi_col = _pick_npi_column(rows)
            if not npi_col:
                raise RuntimeError("Could not identify NPI column.")

            # Create TargetList
            t = TargetList(
                name=name, brand=brand, pharma=pharma,
                indication=indication, notes=notes, filename=(f.filename or None)
            )
            db.session.add(t); db.session.flush()

            # Build unique NPIs + compact 'extra' JSON
            seen = set()
            to_insert = []
            for r in rows:
                npi = _extract_clean_npi(r.get(npi_col))
                if not npi or npi in seen:
                    continue
                seen.add(npi)
                extra = {k: v for k, v in r.items() if k != npi_col}
                if len(extra) > 30:
                    keys = list(extra.keys())[:30]
                    extra = {k: extra[k] for k in keys}
                to_insert.append(TargetListEntry(target_list_id=t.id, npi=npi, extra=extra if extra else None))

            db.session.bulk_save_objects(to_insert)

            # Basic stats
            t.n_rows = int(len(rows))
            t.n_unique_npi = int(len(seen))

            # Coverage (cache or live; safe; now Python-side match for live)
            t.n_matched_network = compute_network_match_count(t.id)

            db.session.commit()
            flash("Target list stored.", "success")
            return redirect(url_for("targets.target_lists"))
        except Exception as e:
            db.session.rollback()
            flash(f"Failed to store target list: {e}", "danger")
            return redirect(url_for("targets.target_lists"))

    lists = db.session.execute(select(TargetList).order_by(TargetList.created_at.desc())).scalars().all()
    return render_template("targets/list.html", lists=lists)

@targets_bp.route("/target-lists/<int:tid>", methods=["GET", "POST"])
def target_list_detail(tid: int):
    try:
        if hasattr(ext, "ensure_mvp_schema_sqlite"):
            ext.ensure_mvp_schema_sqlite()
    except Exception:
        pass

    t = db.session.get(TargetList, tid)
    if not t:
        flash("Target list not found.", "danger")
        return redirect(url_for("targets.target_lists"))

    if request.method == "POST":
        action = request.form.get("action")
        if action == "meta":
            t.name = request.form.get("name") or t.name
            t.brand = request.form.get("brand") or None
            t.pharma = request.form.get("pharma") or None
            t.indication = request.form.get("indication") or None
            t.notes = request.form.get("notes") or None
            db.session.commit()
            flash("Details updated.", "success")
            return redirect(url_for("targets.target_list_detail", tid=tid))
        elif action == "map":
            program_ids = request.form.getlist("program_ids")
            programs = db.session.execute(
                select(Program).where(Program.id.in_([int(x) for x in program_ids])) if program_ids else select(Program).where(False)
            ).scalars().all()
            t.programs = programs
            db.session.commit()
            flash("Mappings updated.", "success")
            return redirect(url_for("targets.target_list_detail", tid=tid))

    q = request.args.get("q", "").strip()
    programs = _query_programs(q)

    # Refresh coverage using the new Python-side match
    t.n_matched_network = compute_network_match_count(t.id)
    db.session.commit()

    facets, numerics, sample = _summarize_entries(t.id)

    return render_template("targets/detail.html", t=t, programs=programs, q=q, facets=facets, numerics=numerics, sample=sample)
